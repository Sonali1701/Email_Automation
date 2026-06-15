"""Create a ready-to-fill contacts.xlsx template with a few example rows.

Run:  python make_sample_contacts.py
Then open contacts.xlsx, delete the example rows, and paste in your real list.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["First Name", "Last Name", "Email", "Title", "Company", "Category Override"]

EXAMPLES = [
    ["Jane", "Doe", "jane.doe@example-hospital.org", "Chief Nursing Officer", "Example Health System", ""],
    ["Mark", "Lee", "mark.lee@example-corp.com", "VP, Talent Acquisition", "Example Corp", ""],
    ["Priya", "Shah", "priya.shah@example-mfg.com", "Director of Procurement", "Example Manufacturing", ""],
    ["Sam", "Roy", "sam.roy@example.com", "Head of Operations", "Example Industries", "hr"],
]

OUT = Path("contacts.xlsx")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill

    for row in EXAMPLES:
        ws.append(row)

    widths = [14, 14, 34, 28, 26, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Keep guidance on a separate sheet so it never pollutes the contact rows.
    info = wb.create_sheet("Instructions")
    lines = [
        "How to use this sheet:",
        "1. Replace the example rows on the 'Contacts' tab with your real list.",
        "2. Email is required. First Name, Title, and Company are recommended.",
        "3. Category Override is optional. Leave blank to let Claude classify the title.",
        "   Valid values: healthcare, hr, procurement, other.",
        "4. Save, then run:  python main.py --dry-run   (preview before sending).",
    ]
    for i, line in enumerate(lines, start=1):
        c = info.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True)
    info.column_dimensions["A"].width = 80

    wb.save(OUT)
    print(f"Wrote {OUT.resolve()} with {len(EXAMPLES)} example rows.")
    print("Open it, replace the examples with your real contacts, then run main.py.")


if __name__ == "__main__":
    main()
